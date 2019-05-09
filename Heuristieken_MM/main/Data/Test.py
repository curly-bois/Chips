import openpyxl as op


book = op.load_workbook("Test.xlsx")
ws = book.worksheets[0]
print(ws.max_row + 1)
for cell in ws["C"]:
    if all is None:
        print(cell.row)
        row = cell.row
        break

else:
    print(cell.row + 1)
