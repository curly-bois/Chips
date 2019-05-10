import openpyxl as op

def make_data(orderlist,not_connected):
    column = 1
    order = f"{orderlist}"
    notcon = f'{not_connected}'
    name1 = "order of points"
    name2 = "not connected points"

    wb = op.load_workbook('Test.xlsx')

    ws = wb.worksheets[0]

    row = ws.max_row + 1


    ws.cell(row=row, column=column).value = name1
    column += 1
    ws.cell(row=row, column=column).value = order
    row = ws.max_row + 1
    column = 1
    ws.cell(row=row, column=column).value = name2
    column += 1
    ws.cell(row=row, column=column).value = notcon

    wb.save('Test.xlsx')
    wb.close()
