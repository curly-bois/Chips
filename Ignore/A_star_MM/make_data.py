import openpyxl as op



def make_data():
    column = 1
    # order = f"{orderlist}"
    # notcon = f'{not_connected}'
    name1 = "order of points"
    name2 = "not connected points"

    wb = op.load_workbook('HeuristiekenData.xlsx')
    # ws = wb.get_sheet_by_name('Sheet1')
    ws = wb.active


    for cell in ws:
        if cell.value is None:
            print( cell.row)
            break
    else:
        print( cell.row + 1)

    # ws.cell(row=row, column=column).value = name1
    # column += 1
    # ws.cell(row=row, column=column).value = order
    # row += 1
    # column = 1
    # ws.cell(row=row, column=column).value = name2
    # column += 1
    # ws.cell(row=row, column=column).value = notcon

    wb.save('HeuristiekenData.xlsx')
    wb.close()
